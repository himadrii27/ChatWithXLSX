import streamlit as st
import openpyxl
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain_openai import ChatOpenAI
from langchain_community.vectorstores import FAISS
from langchain.chains.question_answering import load_qa_chain
from langchain.prompts import PromptTemplate
from langchain.text_splitter import RecursiveCharacterTextSplitter

import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Verify the API key is loaded
if not os.getenv("OPENAI_API_KEY"):
    st.error("OpenAI API key not found. Please check your .env file.")

# Initialize embeddings globally
@st.cache_resource
def get_embeddings():
    return HuggingFaceEmbeddings(
        model_name="sentence-transformers/all-mpnet-base-v2"
    )

def get_excel_text(excel_docs):
    text = ""
    try:
        for excel_file in excel_docs:
            workbook = openpyxl.load_workbook(excel_file)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    text += row_text + "\n"
        return text
    except Exception as e:
        st.error(f"Error reading Excel file: {str(e)}")
        return None

def get_text_chunks(text):
    if not text:
        return None
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=1000,
        chunk_overlap=200
    )
    return text_splitter.split_text(text)

def get_vector_store(text_chunks):
    if not text_chunks:
        return None
    try:
        embeddings = get_embeddings()
        vector_store = FAISS.from_texts(text_chunks, embedding=embeddings)
        return vector_store
    except Exception as e:
        st.error(f"Error creating vector store: {str(e)}")
        return None

def get_conversation_chain():
    prompt_template = """
    Answer the question as detailed as possible from the provided context.
    If the answer is not in the provided context, just say:
    'Answer is not available in the context.'
    Don't provide a wrong or made-up answer.

    Context:
    {context}

    Question:
    {question}

    Answer:
    """
    
    try:
        model = ChatOpenAI(temperature=0.3)
        prompt = PromptTemplate(template=prompt_template, input_variables=["context", "question"])
        chain = load_qa_chain(model, chain_type="stuff", prompt=prompt)
        return chain
    except Exception as e:
        st.error(f"Error creating conversation chain: {str(e)}")
        return None

def process_question(vector_store, question):
    try:
        docs = vector_store.similarity_search(question)
        chain = get_conversation_chain()
        
        if chain:
            response = chain(
                {"input_documents": docs, "question": question},
                return_only_outputs=True
            )
            return response["output_text"]
    except Exception as e:
        st.error(f"Error processing question: {str(e)}")
    return "Sorry, I couldn't process your question. Please try again."

def main():
    st.set_page_config(page_title="Excel Chat Assistant")
    st.title("Chat with Excel Files 📊")

    # Sidebar for file upload
    with st.sidebar:
        st.header("Upload Files")
        excel_docs = st.file_uploader(
            "Upload Excel files",
            type=["xlsx", "xls"],
            accept_multiple_files=True
        )
        process_button = st.button("Process Files")

    # Initialize session state
    if 'vector_store' not in st.session_state:
        st.session_state.vector_store = None

    # Process files when button is clicked
    if process_button and excel_docs:
        with st.spinner("Processing files..."):
            raw_text = get_excel_text(excel_docs)
            if raw_text:
                text_chunks = get_text_chunks(raw_text)
                if text_chunks:
                    vector_store = get_vector_store(text_chunks)
                    if vector_store:
                        st.session_state.vector_store = vector_store
                        st.success("Files processed successfully!")
                    else:
                        st.error("Error creating vector store")
                else:
                    st.error("Error creating text chunks")
            else:
                st.error("Error reading files")

    # Question input
    st.header("Ask Questions")
    question = st.text_input("Enter your question about the Excel files:")

    if question:
        if st.session_state.vector_store is None:
            st.warning("Please upload and process Excel files first!")
        else:
            with st.spinner("Finding answer..."):
                answer = process_question(st.session_state.vector_store, question)
                st.write("Answer:", answer)

if __name__ == "__main__":
    main()